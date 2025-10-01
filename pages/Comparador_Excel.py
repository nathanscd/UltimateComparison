import streamlit as st
import pandas as pd
import time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from rapidfuzz import process, fuzz
from joblib import Parallel, delayed
from difflib import SequenceMatcher

st.set_page_config(page_title="Comparador de Planilhas", page_icon="📊", layout="wide")
st.title("Comparador de Planilhas")
st.markdown("Preencha o formulário inicial para configurar a comparação.")

# === FORMULÁRIO INICIAL ===
with st.form("config_form"):
    n_planilhas = st.number_input("Quantas planilhas você deseja comparar?", min_value=1, max_value=5, value=1, step=1)
    n_colunas = st.number_input("Quantas colunas devem ser comparadas em cada planilha?", min_value=1, max_value=5, value=1, step=1)
    modo = None
    if n_planilhas > 1:
        modo = st.radio("Modo de comparação", ["Primeira planilha contra as outras", "Comparação cruzada entre todas"])
        paralelizar = st.checkbox("Usar processamento paralelo (pode acelerar bastante)", value=True)
        confidence_threshold = st.slider("Limite mínimo de similaridade (%) para considerar um match", 30, 95, 60)
    submitted = st.form_submit_button("Continuar")

    if submitted:
        st.session_state["config_done"] = True
        st.session_state["n_planilhas"] = n_planilhas
        st.session_state["n_colunas"] = n_colunas
        st.session_state["modo"] = modo
        st.session_state["paralelizar"] = paralelizar
        st.session_state["confidence_threshold"] = confidence_threshold
    ## === FINAL DO FORMULÁRIO ===

# === UPLOAD DAS PLANILHAS ===
if st.session_state.get("config_done", False):
    n_planilhas = st.session_state["n_planilhas"]
    n_colunas = st.session_state["n_colunas"]
    modo = st.session_state.get("modo")
    paralelizar = st.session_state.get("paralelizar")
    confidence_threshold = st.session_state.get("confidence_threshold", 60)

    st.subheader("Upload das planilhas")
    arquivos = []
    for i in range(n_planilhas):
        uploaded = st.file_uploader(f"Selecione a planilha {i+1}", type=["xlsx"], key=f"file_{i}")
        arquivos.append(uploaded)

    if all(arquivos):
        planilhas = [pd.ExcelFile(f) for f in arquivos]

        abas = []
        dfs = []
        for i, p in enumerate(planilhas):
            aba = st.selectbox(f"Escolha a aba da planilha {i+1}", p.sheet_names, key=f"aba_{i}")
            abas.append(aba)
            df = p.parse(aba)
            dfs.append(df)

            st.markdown(f"### Pré-visualização da Planilha {i+1} ({aba})")
            st.dataframe(df.head(100))

        st.subheader("Configuração das colunas")
        colunas_escolhidas = []
        for i, df in enumerate(dfs):
            cols = []
            for j in range(n_colunas):
                col = st.selectbox(f"Coluna {j+1} da planilha {i+1}", df.columns, key=f"col_{i}_{j}")
                cols.append(col)
            colunas_escolhidas.append(cols)

        # === Funções auxiliares ===
        def is_empty_like(x):
            return pd.isna(x) or str(x).strip() == "" or str(x).strip() in ["0", "0.0", "-"]

        ## ==== CITAR DIFERENÇAS ====
        def get_diff_summary(a, b):
            a_tokens = str(a).split()
            b_tokens = str(b).split()
            sm = SequenceMatcher(None, a_tokens, b_tokens)
            removed, added, replaced = [], [], []
            for tag, i1, i2, j1, j2 in sm.get_opcodes():
                if tag == "delete":
                    removed.append(" ".join(a_tokens[i1:i2]))
                elif tag == "insert":
                    added.append(" ".join(b_tokens[j1:j2]))
                elif tag == "replace":
                    replaced.append(f"{' '.join(a_tokens[i1:i2])} -> {' '.join(b_tokens[j1:j2])}")
            parts = []
            if removed:
                parts.append("Removido: " + ", ".join(removed))
            if added:
                parts.append("Adicionado: " + ", ".join(added))
            if replaced:
                parts.append("Substituído: " + ", ".join(replaced))
            if not parts:
                return "Nenhuma diferença"
            return "A diferença do primeiro para o segundo é: " + "; ".join(parts)

        ## ==== ENCONTRAR SIMILAR ==== 
        def encontrar_mais_similar_e_diff(texto, lista_textos, threshold=60):
            if is_empty_like(texto):
                return "Nenhuma similaridade encontrada", "Nada encontrado"
            texto = str(texto)
            candidatos = [str(t) for t in lista_textos if not is_empty_like(t)]
            if not candidatos:
                return "Nenhuma similaridade encontrada", "Nada encontrado"
            res = process.extractOne(texto, candidatos, scorer=fuzz.ratio)
            if not res:
                return "Nenhuma similaridade encontrada", "Nada encontrado"
            match, score, _ = res
            if score < threshold:
                return "Nenhuma similaridade encontrada", "Nada encontrado"
            diff_text = get_diff_summary(texto, match)
            return match, diff_text

        # === Botão de iniciar comparação ===
        if st.button("Iniciar Comparação"):
            resultados_finais = []
            total_passos = 0

            if n_planilhas == 1:
                total_passos = len(dfs[0]) * n_colunas
            elif modo == "Primeira planilha contra as outras":
                total_passos = len(dfs[0]) * n_colunas * (n_planilhas - 1)
            else:
                for i in range(n_planilhas):
                    for j in range(i + 1, n_planilhas):
                        total_passos += len(dfs[i]) * n_colunas

            progresso_bar = st.progress(0)
            progresso_text = st.empty()
            progresso_atual = 0
            inicio = time.time()

            if n_planilhas == 1:
                base_df = dfs[0].copy()
                for c in colunas_escolhidas[0]:
                    similares, diffs = [], []
                    valores = base_df[c].tolist()
                    for i, val in enumerate(valores):
                        similar, diff = encontrar_mais_similar_e_diff(val, valores)
                        similares.append(similar)
                        diffs.append(diff)
                        progresso_atual += 1
                        progresso_bar.progress(progresso_atual / total_passos)
                        tempo_passado = time.time() - inicio
                        tempo_estimado = (tempo_passado / progresso_atual) * (total_passos - progresso_atual)
                        tempo_str = f"{tempo_estimado/60:.1f} minutos restantes" if tempo_estimado > 60 else f"{tempo_estimado:.1f} segundos restantes"
                        progresso_text.markdown(f"**Processando {progresso_atual}/{total_passos} itens... | {tempo_str}**")
                    base_df[f"Similar_{c}"] = similares
                    base_df[f"Diferenças_{c}"] = diffs
                resultados_finais.append(("Planilha Única", base_df))

            elif modo == "Primeira planilha contra as outras":
                base_df = dfs[0].copy()
                for idx, df in enumerate(dfs[1:], start=2):
                    for c in colunas_escolhidas[0]:
                        similares, diffs = [], []
                        for i, val in enumerate(base_df[c].tolist()):
                            similar, diff = encontrar_mais_similar_e_diff(val, df[colunas_escolhidas[idx-1][0]])
                            similares.append(similar)
                            diffs.append(diff)
                            progresso_atual += 1
                            progresso_bar.progress(progresso_atual / total_passos)
                            tempo_passado = time.time() - inicio
                            tempo_estimado = (tempo_passado / progresso_atual) * (total_passos - progresso_atual)
                            tempo_str = f"{tempo_estimado/60:.1f} minutos restantes" if tempo_estimado > 60 else f"{tempo_estimado:.1f} segundos restantes"
                            progresso_text.markdown(f"**Processando {progresso_atual}/{total_passos} itens... | {tempo_str}**")
                        base_df[f"Similar_Planilha{idx}_{c}"] = similares
                        base_df[f"Diferenças_Planilha{idx}_{c}"] = diffs
                resultados_finais.append(("Planilha 1 vs Outras", base_df))

            else:
                for i, dfA in enumerate(dfs):
                    for j, dfB in enumerate(dfs):
                        if i < j:
                            temp_df = dfA.copy()
                            for c in colunas_escolhidas[i]:
                                similares, diffs = [], []
                                for val in temp_df[c].tolist():
                                    similar, diff = encontrar_mais_similar_e_diff(val, dfB[colunas_escolhidas[j][0]])
                                    similares.append(similar)
                                    diffs.append(diff)
                                    progresso_atual += 1
                                    progresso_bar.progress(progresso_atual / total_passos)
                                    tempo_passado = time.time() - inicio
                                    tempo_estimado = (tempo_passado / progresso_atual) * (total_passos - progresso_atual)
                                    tempo_str = f"{tempo_estimado/60:.1f} minutos restantes" if tempo_estimado > 60 else f"{tempo_estimado:.1f} segundos restantes"
                                    progresso_text.markdown(f"**Processando {progresso_atual}/{total_passos} itens... | {tempo_str}**")
                                temp_df[f"Similar_Planilha{j+1}_{c}"] = similares
                                temp_df[f"Diferenças_Planilha{j+1}_{c}"] = diffs
                            resultados_finais.append((f"Planilha {i+1} vs {j+1}", temp_df))

            # === Exportar com formatação ===
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                for nome, df in resultados_finais:
                    df.to_excel(writer, sheet_name=nome[:31], index=False)
            output.seek(0)

            wb = load_workbook(output)
            for ws in wb.worksheets:
                for col in ws.columns:
                    col_letter = get_column_letter(col[0].column)
                    ws.column_dimensions[col_letter].width = 50
                    for i, cell in enumerate(col):
                        if i == 0:
                            cell.fill = PatternFill(start_color="ededed", end_color="ededed", fill_type="solid")
                            cell.font = Font(bold=True)
                        else:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')

            output_final = BytesIO()
            wb.save(output_final)
            output_final.seek(0)

            st.success("✅ Comparação concluída com sucesso!")
            st.download_button(
                label="📥 Baixar planilha com diferenças",
                data=output_final,
                file_name="Diferenças.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )